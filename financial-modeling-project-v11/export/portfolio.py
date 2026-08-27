"""Excel export for the educational Markowitz portfolio builder."""
from io import BytesIO


def build_portfolio_excel(result: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio Summary"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    p = result.get("portfolio", {})
    gmv = result.get("minimum_variance_portfolio", {})
    summary = [
        ("Investment amount (KZT, display amount)", result.get("amount_kzt")),
        ("Portfolio base currency", p.get("base_currency", "USD")),
        ("Method", p.get("method")),
        ("Optimization objective", p.get("objective")),
        ("U.S. risk-free rate (% p.a.)", p.get("risk_free_rate_pct")),
        ("Risk-free instrument", (result.get("risk_free") or {}).get("instrument")),
        ("Risk-free rate as of", (result.get("risk_free") or {}).get("as_of")),
        ("Risk-free source", (result.get("risk_free") or {}).get("source")),
        ("Historical model return (%)", p.get("expected_return_pct")),
        ("Historical model gain (KZT)", p.get("expected_gain_kzt")),
        ("Portfolio historical risk σ (%)", p.get("historical_risk_pct")),
        ("Sharpe ratio", p.get("sharpe_ratio")),
        ("Concentration mode", p.get("concentration_mode")),
        ("Maximum position weight (%)", p.get("max_position_weight_pct")),
        ("Portfolio constraints", "Long-only; fully invested; concentration policy depends on selected mode"),
        ("Minimum-variance return (%)", gmv.get("expected_return_pct")),
        ("Minimum-variance risk σ (%)", gmv.get("historical_risk_pct")),
        ("Selected assets", len(result.get("selected", []))),
    ]
    ws.append(["Metric", "Value"])
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
    for row in summary:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28

    wa = wb.create_sheet("Allocation")
    cols = ["Ticker", "Name", "Asset Type", "Market", "Weight %", "Amount KZT", "0% explanation"]
    wa.append(cols)
    for c in wa[1]:
        c.fill = header_fill
        c.font = header_font
    zero_reasons = {x.get("ticker"): x.get("reason") for x in result.get("zero_weight_assets", [])}
    for r in result.get("allocation", []):
        wa.append([
            r.get("ticker"), r.get("name"), r.get("asset_type"), r.get("market"),
            r.get("weight_pct"), r.get("amount_kzt"), zero_reasons.get(r.get("ticker"), ""),
        ])
    for i, w in enumerate([16, 42, 14, 14, 12, 18, 70], 1):
        wa.column_dimensions[get_column_letter(i)].width = w

    wm = wb.create_sheet("Asset Metrics")
    mcols = ["Ticker", "Name", "Type", "Market", "Expected Return %", "Historical Risk σ %", "Observations", "History Method"]
    wm.append(mcols)
    for c in wm[1]:
        c.fill = header_fill
        c.font = header_font
    for r in result.get("individual_metrics", []):
        wm.append([
            r.get("ticker"), r.get("name"), r.get("asset_type"), r.get("market"),
            r.get("expected_return_pct"), r.get("historical_risk_pct"),
            r.get("observations"), r.get("history_method"),
        ])
    for i, w in enumerate([16, 38, 12, 14, 18, 20, 14, 38], 1):
        wm.column_dimensions[get_column_letter(i)].width = w

    def matrix_sheet(title, payload):
        sh = wb.create_sheet(title)
        labels = payload.get("labels", [])
        sh.append([title] + labels)
        for c in sh[1]:
            c.fill = header_fill
            c.font = header_font
        for lab, row in zip(labels, payload.get("matrix", [])):
            sh.append([lab] + row)
        sh.freeze_panes = "B2"
        sh.column_dimensions["A"].width = 16
        for i in range(2, len(labels) + 2):
            sh.column_dimensions[get_column_letter(i)].width = 12
        return sh

    matrix_sheet("Correlation", result.get("correlation", {}))
    matrix_sheet("Covariance", result.get("covariance", {}))

    wf = wb.create_sheet("Efficient Frontier")
    wf.append(["Historical Risk σ %", "Historical Model Return %", "Sharpe"])
    for c in wf[1]:
        c.fill = header_fill
        c.font = header_font
    for pt in result.get("efficient_frontier", {}).get("points", []):
        wf.append([pt.get("risk_pct"), pt.get("return_pct"), pt.get("sharpe")])
    wf.column_dimensions["A"].width = 24
    wf.column_dimensions["B"].width = 28
    wf.column_dimensions["C"].width = 14

    wn = wb.create_sheet("Notes")
    wn["A1"] = "Methodology / limitations"
    wn["A1"].fill = header_fill
    wn["A1"].font = header_font
    row = 2
    for line in result.get("methodology", []):
        wn.cell(row, 1, line)
        row += 1
    for line in result.get("warnings", []):
        wn.cell(row, 1, "Warning: " + line)
        row += 1
    wn.column_dimensions["A"].width = 120
    for r in wn.iter_rows():
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
