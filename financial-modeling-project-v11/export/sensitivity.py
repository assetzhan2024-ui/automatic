"""
export/sensitivity.py
=====================
Dedicated "Sensitivity" sheet for DCF workbook.

Layout (matches NBA_ADVANCED template architecture):

  Row 1   Title
  Row 3   Current market price  (B3 = rec["price"])

  Section A — Perpetuity Growth Method  (cols B..G)
  Row 5   Header "Implied Share Price — Perpetuity Growth Method"
  Row 6   Corner label + TGR column headers
  Rows 7-11  WACC rows × 5 TGR columns → implied share price
  Row 13  Header "% Premium / (Discount) to market — PGM"
  Row 14  Corner + TGR headers
  Rows 15-19  Same grid → (implied - current) / current %

  Section B — Exit EBITDA Multiple Method  (cols I..N)
  Same structure, rows 5-19, but WACC × EBITDA Multiple

  Conditional formatting (CellIsRule):
    implied_price > current_price  →  light green  (#C6EFCE / dark #375623)
    implied_price < current_price  →  light red    (#FFC7CE / dark #9C0006)
    implied_price == current_price →  light yellow (#FFEB9C / dark #9C6500)

Public API:
    build_sensitivity_sheet(wb, rec, fund, suffix="", inp_name="DCF input")
"""

from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    raise ImportError("pip install openpyxl")


# ── Palette (matches NBA template colours) ────────────────────────────────
_BG_DARK    = "1F2D3D"
_BG_HEAD    = "5482AB"    # section header blue (from template row 90)
_BG_SUBHEAD = "165788"    # darker blue (from template row 72/83)
_BG_LABEL   = "EAEAEA"    # light grey for corner/row labels
_BG_CALC    = "FFFFFF"
_BG_BASE    = "E4DFEC"    # purple-ish highlight for base case diagonal
_FG_W       = "FFFFFF"
_FG_D       = "1C2833"
_FG_G       = "717D7E"
_BRD        = "BFC9CA"

# Conditional format fills (Excel standard traffic-light colours)
_CF_GREEN_BG  = "C6EFCE"
_CF_GREEN_FG  = "375623"
_CF_RED_BG    = "FFC7CE"
_CF_RED_FG    = "9C0006"
_CF_YELLOW_BG = "FFEB9C"
_CF_YELLOW_FG = "9C6500"

_FMT_PRICE = "#,##0.00"
_FMT_PCT   = "0.0%"
_FMT_NUM   = "#,##0.0"


def _bd():
    s = Side(style="thin", color=_BRD)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h):
    return PatternFill("solid", fgColor=h)

def _f(sz=9, bold=False, color=_FG_D):
    return Font(name="Calibri", size=sz, bold=bold, color=color)


def _hdr_cell(ws, row, col, text, end_col=None, bg=_BG_HEAD):
    if end_col and end_col > col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", size=10, bold=True, color=_FG_W)
    c.fill      = _fill(bg)
    c.border    = _bd()
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


def _val_cell(ws, row, col, value, bg=_BG_CALC, fmt=None, bold=False,
              is_base=False):
    c = ws.cell(row=row, column=col, value=value)
    actual_bg = _BG_BASE if is_base else bg
    c.font      = Font(name="Calibri", size=9, bold=bold or is_base,
                       color=_FG_D)
    c.fill      = _fill(actual_bg)
    c.border    = _bd()
    c.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        c.number_format = fmt
    return c


# ── Core DCF recalculation ────────────────────────────────────────────────

def _recalc_dcf(rec: dict, fund: Optional[dict],
                wacc: float, tgr: float,
                method: str = "pgm") -> Optional[float]:
    """
    Analytically compute implied share price for given WACC and TGR
    (or EBITDA multiple when method='ebitda').

    Uses the same inputs as _build_dcf_input:
      - revenue base & 10-year projection
      - EBITDA margin (base + convergence)
      - D&A, Capex, NWC % of revenue
      - Tax rate 20%
      - net_debt from balance sheet
      - shares outstanding

    Returns implied price per share (in the ticker's native currency) or None.
    """
    from export.dcf import _get_fund, _safe_ratio, _taper_growth, _converge_margin

    rev_m    = _get_fund(fund, "income", "Total Revenue",  0.0) or 0.0
    ebitda_m = _get_fund(fund, "income", "EBITDA",         None)
    dna_m    = _get_fund(fund, "cashflow",
                         "Depreciation And Amortization",  None)
    capex_m  = _get_fund(fund, "cashflow",
                         "Capital Expenditure",            None)
    debt_m   = _get_fund(fund, "balance", "Total Debt",    0.0) or 0.0
    cash_m   = _get_fund(fund, "balance",
                         "Cash And Cash Equivalents",      0.0) or 0.0
    net_debt = debt_m - cash_m

    ebitda_base = _safe_ratio(ebitda_m, rev_m,
                              default=0.20, min_val=0.0, max_val=0.90)
    dna_pct     = _safe_ratio(abs(dna_m)   if dna_m   else None, rev_m,
                              default=0.03, min_val=0.0, max_val=0.30)
    capex_pct   = _safe_ratio(abs(capex_m) if capex_m else None, rev_m,
                              default=0.05, min_val=0.0, max_val=0.30)
    nwc_pct     = 0.02
    tax         = 0.20
    tv_mult     = {"KZ":6.0,"Emerging":7.0}.get(
                      rec.get("region","US"), 8.0)

    # Shares
    shares = None
    if fund:
        sh = _get_fund(fund, "balance", "Share Issued")
        if sh and sh > 0:
            shares = sh
    if not shares and rec.get("market_cap") and rec.get("price"):
        try:
            shares = (float(rec["market_cap"])
                      / float(rec["price"]) / 1e6)
        except (TypeError, ValueError, ZeroDivisionError):
            shares = 100.0
    shares = shares or 100.0

    if rev_m <= 0 or wacc <= tgr:
        return None

    # Detect recent growth from fund data
    recent_growth = 0.10
    if fund:
        rvs = sorted(
            [(yr, float(v.get("Total Revenue", 0)))
             for yr, v in fund.get("income", {}).items()
             if v.get("Total Revenue")],
            key=lambda x: x[0])
        if len(rvs) >= 2 and rvs[-2][1]:
            recent_growth = max(0.0, min(0.40,
                rvs[-1][1] / rvs[-2][1] - 1))

    growth_schedule = _taper_growth(recent_growth, 10)   # Y+1..Y+10
    margin_schedule = _converge_margin(ebitda_base, 11)  # base..Y+10

    # Project 10 years of UFCF
    pv_fcf = 0.0
    rev = rev_m
    for k in range(1, 11):
        rev      = rev * (1 + growth_schedule[k - 1])
        margin   = margin_schedule[k]
        ebitda   = rev * margin
        dna      = rev * dna_pct
        ebit     = ebitda - dna
        tax_amt  = ebit * tax
        ebiat    = ebit - tax_amt
        capex    = rev * capex_pct
        nwc      = rev * nwc_pct
        ufcf     = ebiat + dna - capex - nwc
        df       = 1.0 / (1.0 + wacc) ** k
        pv_fcf  += ufcf * df

    # Terminal value
    ufcf_final = pv_fcf * 0   # recompute final-year UFCF
    rev_final  = rev_m
    for k in range(1, 11):
        rev_final = rev_final * (1 + growth_schedule[k - 1])
    margin_final = margin_schedule[10]
    ebitda_final = rev_final * margin_final
    ufcf_final   = (ebitda_final * (1 - tax) -
                    rev_final * capex_pct -
                    rev_final * nwc_pct)

    if method == "pgm":
        try:
            tv = ufcf_final * (1 + tgr) / (wacc - tgr)
        except ZeroDivisionError:
            return None
    else:   # exit EBITDA multiple
        tv = ebitda_final * tv_mult

    pv_tv = tv / (1.0 + wacc) ** 10
    ev    = pv_fcf + pv_tv
    eq    = ev - net_debt
    if shares <= 0:
        return None
    return eq / shares   # $ per share


# ── Sheet builder ─────────────────────────────────────────────────────────

def build_sensitivity_sheet(wb, rec: dict, fund: Optional[dict],
                            suffix: str = "",
                            inp_name: str = "DCF input") -> None:
    """
    Create a "Sensitivity{suffix}" sheet in wb.

    Two 5×5 matrices:
      Left  (cols B..G):  WACC × TGR    → implied share price (PGM method)
      Right (cols I..N):  WACC × EBITDA multiple (fixed steps) — PGM only
                          (mirrors template's two-panel layout)

    Plus % premium/(discount) grids below each matrix.
    Conditional formatting applied to both price grids.
    """
    ws = wb.create_sheet(f"Sensitivity{suffix}")

    # Column widths (mirror template proportions)
    for col, w in [
        ("A",3),("B",18),("C",11),("D",11),("E",11),("F",11),("G",11),
        ("H",3),
        ("I",18),("J",11),("K",11),("L",11),("M",11),("N",11),
    ]:
        ws.column_dimensions[col].width = w

    region = rec.get("region", "US")
    price  = rec.get("price")
    name   = rec.get("name") or rec.get("ticker", "X")
    cur    = rec.get("currency") or "USD"

    # Base assumptions (same defaults as _build_dcf_input)
    base_wacc = {"KZ":0.14,"Emerging":0.12,
                 "Europe":0.09,"Asia":0.10}.get(region, 0.094)
    base_tgr  = {"KZ":0.04,"Emerging":0.035}.get(region, 0.025)
    base_mult = {"KZ":6.0, "Emerging":7.0}.get(region, 8.0)

    wacc_range = [base_wacc + d for d in (-0.02, -0.01, 0.0, 0.01, 0.02)]
    tgr_range  = [base_tgr  + d for d in (-0.01, -0.005, 0.0, 0.005, 0.01)]
    mult_range = [base_mult + d for d in (-2.0, -1.0, 0.0, 1.0, 2.0)]

    # ── Row 1: Title ──────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    c1 = ws.cell(row=1, column=1,
                 value=f"Sensitivity Analysis — {name}  [{cur}]")
    c1.font      = Font(name="Calibri", size=12, bold=True, color=_FG_W)
    c1.fill      = _fill(_BG_DARK)
    c1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 3: Current price ──────────────────────────────────────────────
    ws.cell(row=3, column=2,
            value="Current market price").font = _f(9, True, _FG_D)
    c_price = ws.cell(row=3, column=3,
                      value=round(float(price), 4) if price else None)
    c_price.font         = Font(name="Calibri", size=10, bold=True,
                                color="1A5276")
    c_price.fill         = _fill("D6EAF8")
    c_price.border       = _bd()
    c_price.number_format = _FMT_PRICE
    c_price.alignment    = Alignment(horizontal="right", vertical="center")

    # ── Helper: write a 5×5 grid ─────────────────────────────────────────
    def _write_grid(start_row: int, start_col: int,
                    col_vals: list, row_vals: list,
                    col_label: str, row_label: str,
                    col_fmt: str, row_fmt: str,
                    method: str):
        """
        Write corner label, column headers, row labels, and 5×5 values.
        Returns list of (row, col) cell addresses of the value cells.
        """
        # Corner
        corner = ws.cell(row=start_row, column=start_col,
                         value=f"{row_label} \\ {col_label}")
        corner.font      = _f(9, True, _FG_D)
        corner.fill      = _fill(_BG_LABEL)
        corner.border    = _bd()
        corner.alignment = Alignment(horizontal="center", vertical="center")

        # Column headers (TGR or multiple)
        for j, cv in enumerate(col_vals):
            c = ws.cell(row=start_row, column=start_col + 1 + j,
                        value=cv)
            c.font          = Font(name="Calibri", size=9, bold=True,
                                   color=_FG_W)
            c.fill          = _fill(_BG_SUBHEAD)
            c.border        = _bd()
            c.number_format = col_fmt
            c.alignment     = Alignment(horizontal="center",
                                        vertical="center")

        addrs = []
        for i, w_v in enumerate(row_vals):
            row = start_row + 1 + i
            # Row label (WACC)
            rl = ws.cell(row=row, column=start_col, value=w_v)
            rl.font          = Font(name="Calibri", size=9, bold=True,
                                    color=_FG_W)
            rl.fill          = _fill(_BG_SUBHEAD)
            rl.border        = _bd()
            rl.number_format = row_fmt
            rl.alignment     = Alignment(horizontal="center",
                                         vertical="center")

            for j, cv in enumerate(col_vals):
                col      = start_col + 1 + j
                is_base  = (abs(w_v - base_wacc) < 1e-9 and
                            abs(cv  - (base_tgr if method == "pgm"
                                       else base_mult)) < 1e-9)
                if method == "pgm":
                    val = _recalc_dcf(rec, fund, w_v, cv, "pgm")
                else:
                    val = _recalc_dcf_mult(rec, fund, w_v, cv)

                txt = round(val, 2) if val is not None else "N/A"
                _val_cell(ws, row, col, txt,
                          fmt=_FMT_PRICE,
                          is_base=is_base)
                addrs.append(get_column_letter(col) + str(row))

        return addrs

    # ── Section A: PGM  (cols B..G) ──────────────────────────────────────
    _hdr_cell(ws, 5, 2,
              "Implied Share Price — Perpetuity Growth Method  (WACC × TGR)",
              end_col=7, bg=_BG_HEAD)

    pgm_addrs = _write_grid(
        start_row=6, start_col=2,
        col_vals=tgr_range,   col_label="TGR",  col_fmt=_FMT_PCT,
        row_vals=wacc_range,  row_label="WACC", row_fmt=_FMT_PCT,
        method="pgm",
    )

    # % premium below PGM grid (rows 12-17)
    _hdr_cell(ws, 12, 2,
              "% Premium / (Discount) to Market  — PGM",
              end_col=7, bg=_BG_SUBHEAD)
    corner2 = ws.cell(row=13, column=2, value="WACC \\ TGR")
    corner2.font = _f(9, True, _FG_D)
    corner2.fill = _fill(_BG_LABEL); corner2.border = _bd()
    for j, cv in enumerate(tgr_range):
        c = ws.cell(row=13, column=3+j, value=cv)
        c.font = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        c.fill = _fill(_BG_SUBHEAD); c.border = _bd()
        c.number_format = _FMT_PCT
    for i, w_v in enumerate(wacc_range):
        row = 14 + i
        rl = ws.cell(row=row, column=2, value=w_v)
        rl.font = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        rl.fill = _fill(_BG_SUBHEAD); rl.border = _bd()
        rl.number_format = _FMT_PCT
        for j, cv in enumerate(tgr_range):
            col    = 3 + j
            val    = _recalc_dcf(rec, fund, w_v, cv, "pgm")
            if val is not None and price:
                prem = (val - float(price)) / float(price)
                c = ws.cell(row=row, column=col, value=round(prem, 4))
                c.number_format = _FMT_PCT
            else:
                c = ws.cell(row=row, column=col, value="N/A")
            is_base = (abs(w_v - base_wacc) < 1e-9 and
                       abs(cv - base_tgr)   < 1e-9)
            c.font   = _f(9, is_base, _FG_D)
            c.fill   = _fill(_BG_BASE if is_base else _BG_CALC)
            c.border = _bd()
            c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Section B: Exit EBITDA Multiple  (cols I..N) ─────────────────────
    _hdr_cell(ws, 5, 9,
              "Implied Share Price — Exit EBITDA Multiple  (WACC × Multiple)",
              end_col=14, bg=_BG_HEAD)

    mult_addrs = _write_grid(
        start_row=6, start_col=9,
        col_vals=mult_range,  col_label="EV/EBITDA",  col_fmt="0.0\"x\"",
        row_vals=wacc_range,  row_label="WACC",        row_fmt=_FMT_PCT,
        method="mult",
    )

    # % premium below multiple grid
    _hdr_cell(ws, 12, 9,
              "% Premium / (Discount) to Market  — EBITDA Multiple",
              end_col=14, bg=_BG_SUBHEAD)
    corner3 = ws.cell(row=13, column=9, value="WACC \\ EV/EBITDA")
    corner3.font = _f(9, True, _FG_D)
    corner3.fill = _fill(_BG_LABEL); corner3.border = _bd()
    for j, cv in enumerate(mult_range):
        c = ws.cell(row=13, column=10+j, value=cv)
        c.font = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        c.fill = _fill(_BG_SUBHEAD); c.border = _bd()
        c.number_format = "0.0\"x\""
    for i, w_v in enumerate(wacc_range):
        row = 14 + i
        rl = ws.cell(row=row, column=9, value=w_v)
        rl.font = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        rl.fill = _fill(_BG_SUBHEAD); rl.border = _bd()
        rl.number_format = _FMT_PCT
        for j, cv in enumerate(mult_range):
            col  = 10 + j
            val  = _recalc_dcf_mult(rec, fund, w_v, cv)
            if val is not None and price:
                prem = (val - float(price)) / float(price)
                c = ws.cell(row=row, column=col, value=round(prem, 4))
                c.number_format = _FMT_PCT
            else:
                c = ws.cell(row=row, column=col, value="N/A")
            is_base = (abs(w_v - base_wacc) < 1e-9 and
                       abs(cv - base_mult)   < 1e-9)
            c.font   = _f(9, is_base, _FG_D)
            c.fill   = _fill(_BG_BASE if is_base else _BG_CALC)
            c.border = _bd()
            c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Conditional formatting on price grids ─────────────────────────────
    if price:
        cp = float(price)
        # green  = above current price
        # red    = below current price
        # yellow = within ±1% of current price (near market)
        green_fill  = PatternFill("solid", fgColor=_CF_GREEN_BG)
        red_fill    = PatternFill("solid", fgColor=_CF_RED_BG)
        yellow_fill = PatternFill("solid", fgColor=_CF_YELLOW_BG)
        green_font  = Font(name="Calibri", size=9, color=_CF_GREEN_FG)
        red_font    = Font(name="Calibri", size=9, color=_CF_RED_FG)
        yellow_font = Font(name="Calibri", size=9, color=_CF_YELLOW_FG)

        near_lo = round(cp * 0.99, 4)
        near_hi = round(cp * 1.01, 4)

        for cell_range in ["C7:G11", "J7:N11"]:
            # Order matters: more specific rules first
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="between",
                           formula=[str(near_lo), str(near_hi)],
                           fill=yellow_fill, font=yellow_font))
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="greaterThan",
                           formula=[str(near_hi)],
                           fill=green_fill, font=green_font))
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="lessThan",
                           formula=[str(near_lo)],
                           fill=red_fill, font=red_font))

    # ── Legend ────────────────────────────────────────────────────────────
    ws.merge_cells("B21:N21")
    leg = ws.cell(
        row=21, column=2,
        value=("🟢 Above market price  |  🔴 Below market price  "
               "|  🟡 Within ±1% of market  |  "
               "Shaded = base case  |  Values = implied share price"))
    leg.font = Font(name="Calibri", size=8, italic=True, color=_FG_G)


def _recalc_dcf_mult(rec: dict, fund: Optional[dict],
                     wacc: float, tv_multiple: float) -> Optional[float]:
    """
    Compute implied share price using exit EBITDA multiple TV method.
    Mirrors _recalc_dcf but uses tv_multiple instead of Gordon Growth.
    """
    from export.dcf import (_get_fund, _safe_ratio,
                             _taper_growth, _converge_margin)

    rev_m    = _get_fund(fund, "income", "Total Revenue",  0.0) or 0.0
    ebitda_m = _get_fund(fund, "income", "EBITDA",         None)
    dna_m    = _get_fund(fund, "cashflow",
                         "Depreciation And Amortization",  None)
    capex_m  = _get_fund(fund, "cashflow",
                         "Capital Expenditure",            None)
    debt_m   = _get_fund(fund, "balance", "Total Debt",    0.0) or 0.0
    cash_m   = _get_fund(fund, "balance",
                         "Cash And Cash Equivalents",      0.0) or 0.0
    net_debt = debt_m - cash_m

    ebitda_base = _safe_ratio(ebitda_m, rev_m,
                              default=0.20, min_val=0.0, max_val=0.90)
    dna_pct     = _safe_ratio(abs(dna_m)   if dna_m   else None, rev_m,
                              default=0.03, min_val=0.0, max_val=0.30)
    capex_pct   = _safe_ratio(abs(capex_m) if capex_m else None, rev_m,
                              default=0.05, min_val=0.0, max_val=0.30)
    nwc_pct = 0.02
    tax     = 0.20

    shares = None
    if fund:
        sh = _get_fund(fund, "balance", "Share Issued")
        if sh and sh > 0:
            shares = sh
    if not shares and rec.get("market_cap") and rec.get("price"):
        try:
            shares = (float(rec["market_cap"])
                      / float(rec["price"]) / 1e6)
        except (TypeError, ValueError, ZeroDivisionError):
            shares = 100.0
    shares = shares or 100.0

    if rev_m <= 0:
        return None

    recent_growth = 0.10
    if fund:
        rvs = sorted(
            [(yr, float(v.get("Total Revenue", 0)))
             for yr, v in fund.get("income", {}).items()
             if v.get("Total Revenue")],
            key=lambda x: x[0])
        if len(rvs) >= 2 and rvs[-2][1]:
            recent_growth = max(0.0, min(0.40,
                rvs[-1][1] / rvs[-2][1] - 1))

    growth_schedule = _taper_growth(recent_growth, 10)
    margin_schedule = _converge_margin(ebitda_base, 11)

    pv_fcf = 0.0
    rev    = rev_m
    for k in range(1, 11):
        rev     = rev * (1 + growth_schedule[k - 1])
        margin  = margin_schedule[k]
        ebitda  = rev * margin
        dna     = rev * dna_pct
        ebit    = ebitda - dna
        ebiat   = ebit * (1 - tax)
        capex   = rev * capex_pct
        nwc     = rev * nwc_pct
        ufcf    = ebiat + dna - capex - nwc
        df      = 1.0 / (1.0 + wacc) ** k
        pv_fcf += ufcf * df

    # Final year EBITDA for TV
    rev_final    = rev_m
    for k in range(1, 11):
        rev_final = rev_final * (1 + growth_schedule[k - 1])
    ebitda_final = rev_final * margin_schedule[10]
    tv           = ebitda_final * tv_multiple
    pv_tv        = tv / (1.0 + wacc) ** 10

    ev = pv_fcf + pv_tv
    eq = ev - net_debt
    if shares <= 0:
        return None
    return eq / shares
