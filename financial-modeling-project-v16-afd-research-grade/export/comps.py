"""
export/comps.py
===============
Trading Comparables Analysis sheet.

Layout:
  Row 1   Title
  Row 3   Legend (colour scale explanation)
  Row 5   Column headers: Ticker | Name | Sector | Region | …metrics…
  Row 6+  One row per company, coloured by quartile vs group median

Colour scale for multiples (lower = cheaper = better):
  <= Q1  →  dark green   "cheap"
  <= med →  light green
  <= Q3  →  orange
  >  Q3  →  red          "expensive"

Mirror logic for profitability (ROE%, Net Margin%, ROA%) — higher = better:
  >= Q3  →  dark green
  >= med →  light green
  >= Q1  →  orange
  <  Q1  →  red

Also writes a CCA sheet identical to export/excel.py._cca_sheet but now with
the full quartile colour logic applied.

Public API:
    build_comps_sheet(wb, records)  — adds "Comps" sheet to existing wb
    build_comps_excel(records)      — standalone workbook → bytes
"""

import io
import statistics
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("pip install openpyxl")


# ── Palette ────────────────────────────────────────────────────────────────
_BG_DARK    = "1F2D3D"
_BG_HEAD    = "2C3E50"
_BG_ALT     = "F4F6F7"
_BG_VALUE   = "FFFFFF"
_FG_W       = "FFFFFF"
_FG_D       = "1C2833"
_FG_G       = "717D7E"
_BRD        = "BFC9CA"

# Quartile traffic-light colours
_DK_GREEN   = ("1E5631", "D5F5E3")   # (font, bg) — cheap / high profitability
_LT_GREEN   = ("1E8449", "EAFAF1")
_ORANGE     = ("7D6608", "FDFDE7")   # using warm yellow instead of orange bg
_RED        = ("C0392B", "FADBD8")
_NEUTRAL    = (_FG_G,    _BG_VALUE)

_FMT_PRICE  = "#,##0.00"
_FMT_PCT    = "0.0%"
_FMT_NUM    = "#,##0.00"
_FMT_BIG    = "#,##0.0"


def _bd():
    s = Side(style="thin", color=_BRD)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)
def _f(sz=9, bold=False, color=_FG_D):
    return Font(name="Calibri", size=sz, bold=bold, color=color)


# ── Statistics helpers ────────────────────────────────────────────────────

def _percentile(vals: list[float], p: float) -> float:
    """Simple linear interpolation percentile (0-100 scale)."""
    if not vals:
        return 0.0
    s = sorted(vals)
    n = len(s)
    if n == 1:
        return s[0]
    idx = p / 100.0 * (n - 1)
    lo  = int(idx)
    hi  = min(lo + 1, n - 1)
    return s[lo] + (s[hi] - s[lo]) * (idx - lo)


def _quartile_color(value: float, q1: float, med: float, q3: float,
                    higher_is_better: bool) -> tuple[str, str]:
    """
    Return (font_hex, bg_hex) for a value given distribution quartiles.
    higher_is_better=False  → multiples (P/E, EV/EBITDA …)
    higher_is_better=True   → profitability (ROE%, Net Margin%, ROA%)
    """
    if higher_is_better:
        if   value >= q3:  return _DK_GREEN
        elif value >= med: return _LT_GREEN
        elif value >= q1:  return _ORANGE
        else:              return _RED
    else:
        if   value <= q1:  return _DK_GREEN
        elif value <= med: return _LT_GREEN
        elif value <= q3:  return _ORANGE
        else:              return _RED


# ── Column definitions ────────────────────────────────────────────────────
# (field_key, display_label, higher_is_better, fmt, width)
_COLS = [
    ("ticker",          "Ticker",      None,  None,      10),
    ("name",            "Name",        None,  None,      22),
    ("sector",          "Sector",      None,  None,      16),
    ("region",          "Region",      None,  None,      10),
    ("currency",        "CCY",         None,  None,       6),
    ("price_usd",       "Price (USD)",  None, _FMT_PRICE, 11),
    ("market_cap_usd",  "Mkt Cap",      None, _FMT_BIG,   13),
    # Multiples — lower is better
    ("pe_ratio",        "P/E",         False, _FMT_NUM,    8),
    ("pb_ratio",        "P/B",         False, _FMT_NUM,    8),
    ("ps_ratio",        "P/S",         False, _FMT_NUM,    8),
    ("ev_ebitda",       "EV/EBITDA",   False, _FMT_NUM,   10),
    ("ev_revenue",      "EV/Rev",      False, _FMT_NUM,    9),
    ("de_ratio",        "D/E",         False, _FMT_NUM,    8),
    # Profitability — higher is better
    ("roe_pct",         "ROE %",       True,  _FMT_PCT,    8),
    ("roa_pct",         "ROA %",       True,  _FMT_PCT,    8),
    ("net_margin",      "Net Mgn %",   True,  _FMT_PCT,    9),
    # Other
    ("eps_trailing",    "EPS",         None,  _FMT_NUM,    9),
    ("score_pct",       "Score",       True,  None,        7),
]

# Fields that need to be coloured
_COLOUR_FIELDS = {d[0] for d in _COLS if d[2] is not None}


def _derive_net_margin(rec: dict) -> Optional[float]:
    """Net income / revenue — derived if not in record."""
    ni  = rec.get("net_income_usd") or rec.get("net_income")
    rev = rec.get("revenue_usd")    or rec.get("revenue")
    # Fallback: infer revenue from market_cap * P/S
    if not rev and rec.get("ps_ratio") and rec.get("market_cap_usd"):
        try:
            rev = float(rec["market_cap_usd"]) / float(rec["ps_ratio"])
        except (TypeError, ValueError, ZeroDivisionError):
            pass
    if ni and rev:
        try:
            return float(ni) / float(rev)
        except (TypeError, ValueError, ZeroDivisionError):
            pass
    return None


def _enrich(rec: dict) -> dict:
    """Add derived fields to record dict (non-destructive copy)."""
    r = dict(rec)
    r.setdefault("net_margin", _derive_net_margin(rec))
    # Use USD price if available
    if r.get("price_usd") is None:
        r["price_usd"] = r.get("price")
    if r.get("market_cap_usd") is None:
        r["market_cap_usd"] = r.get("market_cap")
    return r


def _fmt_val(val, field: str, fmt: Optional[str]) -> str:
    """Format a value for display (before writing to cell)."""
    if val is None:
        return "—"
    if field in ("market_cap_usd",):
        # Format as $B / $M
        try:
            v = float(val)
            if   abs(v) >= 1e12: return f"${v/1e12:.2f}T"
            elif abs(v) >= 1e9:  return f"${v/1e9:.2f}B"
            elif abs(v) >= 1e6:  return f"${v/1e6:.2f}M"
            return f"${v:.0f}"
        except (TypeError, ValueError):
            return str(val)
    if field == "score_pct":
        try:   return f"{int(val)}%"
        except: return str(val)
    return val   # let Excel format via number_format


def build_comps_sheet(wb, records: list[dict]) -> None:
    """Add a 'Comps' sheet to the given workbook."""
    ws = wb.create_sheet("Comps")
    _write_comps(ws, records)


def build_comps_excel(records: list[dict]) -> bytes:
    """Build a standalone workbook with a Comps sheet. Returns bytes."""
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = wb.create_sheet("Comps")
    _write_comps(ws, records)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_comps(ws, records: list[dict]) -> None:
    """Write the full Comps sheet content."""

    enriched = [_enrich(r) for r in records]

    # ── Pre-compute quartiles for each coloured field ─────────────────────
    quartiles: dict[str, tuple] = {}
    for field, _, hib, _, _ in _COLS:
        if hib is None:
            continue
        vals = []
        for r in enriched:
            v = r.get(field)
            if v is not None:
                try:   vals.append(float(v))
                except: pass
        if vals:
            q1  = _percentile(vals, 25)
            med = _percentile(vals, 50)
            q3  = _percentile(vals, 75)
            quartiles[field] = (q1, med, q3)

    # ── Column widths ─────────────────────────────────────────────────────
    for i, (_, _, _, _, w) in enumerate(_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: Title ──────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(_COLS))}1")
    c1 = ws.cell(row=1, column=1,
                 value="Trading Comparables Analysis")
    c1.font      = Font(name="Calibri", size=12, bold=True, color=_FG_W)
    c1.fill      = _fill(_BG_DARK)
    c1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2: Legend ─────────────────────────────────────────────────────
    legend_parts = [
        ("■ Q1 (cheap/best)",   _DK_GREEN[1], _DK_GREEN[0]),
        ("■ Median",            _LT_GREEN[1], _LT_GREEN[0]),
        ("■ Q3",                _ORANGE[1],   _ORANGE[0]),
        ("■ >Q3 (exp/worst)",   _RED[1],      _RED[0]),
    ]
    for i, (txt, bg, fg) in enumerate(legend_parts):
        c = ws.cell(row=2, column=1 + i * 2, value=txt)
        c.font      = Font(name="Calibri", size=8, bold=True, color=fg)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 3: Quartile stats header ──────────────────────────────────────
    # Show Q1 / Median / Q3 for each coloured column
    ws.cell(row=3, column=1, value="Q1 / Median / Q3"
            ).font = _f(8, True, _FG_G)
    for i, (field, label, hib, fmt, _) in enumerate(_COLS, 1):
        if hib is None or field not in quartiles:
            continue
        q1, med, q3 = quartiles[field]
        def _fv(v): return f"{v:.1%}" if fmt == _FMT_PCT else f"{v:.2f}"
        c = ws.cell(row=3, column=i,
                    value=f"{_fv(q1)} / {_fv(med)} / {_fv(q3)}")
        c.font      = _f(8, False, _FG_G)
        c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Row 4: Column headers ─────────────────────────────────────────────
    ws.row_dimensions[4].height = 26
    for i, (_, label, _, _, _) in enumerate(_COLS, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font      = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        c.fill      = _fill(_BG_HEAD)
        c.border    = _bd()
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.freeze_panes = "A5"

    # ── Rows 5+: Data ─────────────────────────────────────────────────────
    for ri, rec in enumerate(enriched, 5):
        alt = ri % 2 == 0
        bg  = _BG_ALT if alt else _BG_VALUE

        for ci, (field, _, hib, fmt, _) in enumerate(_COLS, 1):
            val     = rec.get(field)
            display = _fmt_val(val, field, fmt)

            # Determine cell colours
            if hib is not None and val is not None and field in quartiles:
                try:
                    fv = float(val)
                    q1, med, q3 = quartiles[field]
                    fg_hex, bg_hex = _quartile_color(fv, q1, med, q3, hib)
                except (TypeError, ValueError):
                    fg_hex, bg_hex = _NEUTRAL
            else:
                fg_hex, bg_hex = _FG_D, bg

            c = ws.cell(row=ri, column=ci,
                        value=display if isinstance(display, str)
                              else val)
            c.font      = Font(name="Calibri", size=9, color=fg_hex)
            c.fill      = _fill(bg_hex)
            c.border    = _bd()
            c.alignment = Alignment(
                horizontal="left" if ci <= 4 else "right",
                vertical="center")
            if fmt and not isinstance(display, str):
                c.number_format = fmt

    # ── Median row ────────────────────────────────────────────────────────
    med_row = 5 + len(enriched)
    ws.cell(row=med_row, column=1, value="Sector Median"
            ).font = Font(name="Calibri", size=9, bold=True, color=_FG_W)
    ws.cell(row=med_row, column=1).fill = _fill(_BG_DARK)
    ws.cell(row=med_row, column=2, value="—"
            ).font = _f(9, False, _FG_G)

    for ci, (field, _, hib, fmt, _) in enumerate(_COLS, 1):
        if hib is None or field not in quartiles:
            continue
        _, med, _ = quartiles[field]
        c = ws.cell(row=med_row, column=ci, value=med)
        c.font   = Font(name="Calibri", size=9, bold=True, color=_FG_W)
        c.fill   = _fill(_BG_DARK)
        c.border = _bd()
        c.alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            c.number_format = fmt
